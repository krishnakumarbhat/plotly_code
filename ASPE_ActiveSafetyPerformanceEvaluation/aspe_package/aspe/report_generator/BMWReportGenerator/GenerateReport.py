import pickle
import os
import numpy as np
import openpyxl
from openpyxl.workbook import Workbook

from aspe.report_generator.BMWReportGenerator import BMW_ASPE_SIGNAL_MAPPER


class BMW_report_generator():
    fields = [
        '1-sigma',
        'Bias',
        'max error',
        'latency',
        'NEES-Test',
        'mean',
        '1-sigma',
        'Bias',
        'max error',
        'latency',
        'NEES-Test',
        'mean',
        'NEES-Test',
        'mean',
        '1-sigma',
        'max error',
        'NEES-Test',
        'mean',
        'max error',
        '1-sigma',
        'NEES-Test',
        'mean',
        '1-sigma',
        'Bias',
        'max error',
        'latency',
        'NEES-Test',
        'mean',
        '1-sigma',
        'Bias',
        'max error',
        'latency',
        'NEES-Test',
        'mean',
        '1-sigma',
        'Bias',
        'max error',
        'latency',
        'NEES-Test',
        'mean',
        '1-sigma',
        'Bias',
        'max error',
        'latency',
        'NEES-Test',
        'mean',
        'NEES-Test',
        'mean',
        '1-sigma',
        'Bias',
        'max error',
        'latency',
        'NEES-Test',
        'mean',
        '1-sigma',
        'Bias',
        'max error',
        'latency',
        'NEES-Test',
        'mean',
        '1-sigma',
        'Bias',
        'max error',
        'latency',
        'NEES-Test',
        'mean',
        '1-sigma',
        'Bias',
        'max error',
        'latency',
        'NEES-Test',
        'mean',
        'NEES-Test',
        'mean',
        '1-sigma',
        'Bias',
        'max error',
        'latency',
        'NEES-Test',
        'mean',
        '1-sigma',
        'Bias',
        'max error',
        'latency',
        'NEES-Test',
        'mean',
        'TruePositiveRate',
        'PositivePredictiveValue',
        'correct-classification-rate',
    ]


    signal_mapper_stat = {
        'default':  {
            '1-sigma':   'STD',
            'Bias':      'bias',
            'max error': 'max'},
        'extended': {
            'mean':      'mean',
            'NEES-Test': 'NEES mean Gamma test p-value'},
        'binary':{
            'TruePositiveRate': 'TPR',
            'PositivePredictiveValue': 'PPV'
        }

    }

    def __init__(self, path = "", output_path="example.xlsx", signal_mapper_feature=BMW_ASPE_SIGNAL_MAPPER):
        self.signal_mapper_feature = signal_mapper_feature
        self.path = path
        self.output_path = output_path
        self.book = Workbook()
        self.ws = self.book.active


    def initialize_structure(self):
        if self.path == "":
            self.ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=5)
            self.ws.cell(row=1, column=2).value = 'x-position of reference point'
            self.ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=7)
            self.ws.cell(row=1, column=6).value = 'std x'
            self.ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=11)
            self.ws.cell(row=1, column=8).value = 'y-position of reference point'
            self.ws.merge_cells(start_row=1, start_column=12, end_row=1, end_column=13)
            self.ws.cell(row=1, column=12).value = 'std y'
            self.ws.merge_cells(start_row=1, start_column=14, end_row=1, end_column=15)
            self.ws.cell(row=1, column=14).value = 'covariance_x_y'
            self.ws.merge_cells(start_row=1, start_column=16, end_row=1, end_column=17)
            self.ws.cell(row=1, column=16).value = 'length'
            self.ws.merge_cells(start_row=1, start_column=18, end_row=1, end_column=19)
            self.ws.cell(row=1, column=18).value = 'std length'
            self.ws.merge_cells(start_row=1, start_column=20, end_row=1, end_column=21)
            self.ws.cell(row=1, column=20).value = 'width'
            self.ws.merge_cells(start_row=1, start_column=22, end_row=1, end_column=23)
            self.ws.cell(row=1, column=22).value = 'std width'
            self.ws.merge_cells(start_row=1, start_column=24, end_row=1, end_column=27)
            self.ws.cell(row=1, column=24).value = 'vx_absolute'
            self.ws.merge_cells(start_row=1, start_column=28, end_row=1, end_column=29)
            self.ws.cell(row=1, column=28).value = 'std vx abs'
            self.ws.merge_cells(start_row=1, start_column=30, end_row=1, end_column=33)
            self.ws.cell(row=1, column=30).value = 'vx_relative'
            self.ws.merge_cells(start_row=1, start_column=34, end_row=1, end_column=35)
            self.ws.cell(row=1, column=34).value = 'std vx rel'
            self.ws.merge_cells(start_row=1, start_column=36, end_row=1, end_column=39)
            self.ws.cell(row=1, column=36).value = 'vy_absolute'
            self.ws.merge_cells(start_row=1, start_column=40, end_row=1, end_column=41)
            self.ws.cell(row=1, column=40).value = 'std vy abs'
            self.ws.merge_cells(start_row=1, start_column=42, end_row=1, end_column=45)
            self.ws.cell(row=1, column=42).value = 'vy_relative'
            self.ws.merge_cells(start_row=1, start_column=46, end_row=1, end_column=47)
            self.ws.cell(row=1, column=46).value = 'std vy rel'
            self.ws.merge_cells(start_row=1, start_column=48, end_row=1, end_column=49)
            self.ws.cell(row=1, column=48).value = 'covariance_vx_vy'
            self.ws.merge_cells(start_row=1, start_column=50, end_row=1, end_column=53)
            self.ws.cell(row=1, column=50).value = 'ax_absolute'
            self.ws.merge_cells(start_row=1, start_column=54, end_row=1, end_column=55)
            self.ws.cell(row=1, column=54).value = 'std ax abs'
            self.ws.merge_cells(start_row=1, start_column=56, end_row=1, end_column=59)
            self.ws.cell(row=1, column=56).value = 'ax_relative'
            self.ws.merge_cells(start_row=1, start_column=60, end_row=1, end_column=61)
            self.ws.cell(row=1, column=60).value = 'std ax rel'
            self.ws.merge_cells(start_row=1, start_column=62, end_row=1, end_column=65)
            self.ws.cell(row=1, column=62).value = 'ay_absolute'
            self.ws.merge_cells(start_row=1, start_column=66, end_row=1, end_column=67)
            self.ws.cell(row=1, column=66).value = 'std ay abs'
            self.ws.merge_cells(start_row=1, start_column=68, end_row=1, end_column=71)
            self.ws.cell(row=1, column=68).value = 'ay_relative '
            self.ws.merge_cells(start_row=1, start_column=72, end_row=1, end_column=73)
            self.ws.cell(row=1, column=72).value = 'std ay rel'
            self.ws.merge_cells(start_row=1, start_column=74, end_row=1, end_column=75)
            self.ws.cell(row=1, column=74).value = 'covariance_ax_ay'
            self.ws.merge_cells(start_row=1, start_column=76, end_row=1, end_column=79)
            self.ws.cell(row=1, column=76).value = 'orientation angle'
            self.ws.merge_cells(start_row=1, start_column=80, end_row=1, end_column=81)
            self.ws.cell(row=1, column=80).value = 'std orientation'
            self.ws.merge_cells(start_row=1, start_column=82, end_row=1, end_column=85)
            self.ws.cell(row=1, column=82).value = 'change rate of orientation angle'
            self.ws.merge_cells(start_row=1, start_column=86, end_row=1, end_column=87)
            self.ws.cell(row=1, column=86).value = 'std yaw rate'
            self.ws.merge_cells(start_row=1, start_column=88, end_row=1, end_column=89)
            self.ws.cell(row=1, column=88).value = 'object detection rate'
            self.ws.cell(row=1, column=90).value = 'type/class'

            def as_text(value):
                if value is None:
                    return ""
                return str(value)

            for column, text in enumerate(self.fields, start=2):
                self.ws.cell(column=column, row=2, value=text)

            for column_cells in self.ws.columns:
                length = max(len(as_text(cell.value)) for cell in column_cells)
                self.ws.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length
        else:
            self.book = openpyxl.load_workbook(self.path)
            self.ws = self.book.active

    def write_to_sheet(self, data, name):
        stats = self.ws[2]
        evaluated_features = data.kpis_pairs_features_aggregated
        binary_results = data.kpis_binary_class_aggregated
        kpis = []
        for merged in self.ws.merged_cell_ranges:
            feature = self.ws.cell(column=merged.left[0][1], row=merged.left[0][0]).value

            for index in range(merged.bounds[0], merged.bounds[2] + 1):
                stat = self.ws.cell(column=index, row=2).value
                try:
                    if feature == 'object detection rate':
                        stat_type = 'binary'
                        selected = binary_results[binary_results.signature == self.signal_mapper_stat[stat_type][stat]].reset_index()
                        result = selected['value'][0]
                    else:
                        stat_type = 'extended' if 'std' in feature or 'cov' in feature else 'default'
                        selected = evaluated_features[
                            np.logical_and(evaluated_features.feature_signature == self.signal_mapper_feature[feature],
                                           evaluated_features.kpi_signature == self.signal_mapper_stat[stat_type][stat])].reset_index()

                        result  = selected['kpi_value'][0]
                    kpis.append(result)

                except KeyError as e:
                    print(f"Statistic {feature} - {stat} is missing")
                    kpis.append(None)

                except IndexError as e:
                    print(e)
                    kpis.append(None)

        max_row = self.ws.max_row
        self.ws.cell(column=1, row=max_row+1, value=name)
        for column, text in enumerate(kpis, start=2):
            self.ws.cell(column=column, row=max_row + 1, value=text)

    def save_workbook(self):
        self.book.save(self.output_path)


if __name__ == "__main__":
    input_path = r"C:\Users\zj9lvp\Documents\aspe_bmw_data\A410\results"
    report = BMW_report_generator(output_path=r"C:\Users\zj9lvp\Documents\aspe_bmw_data\A410\results\output_2.xlsx", )
    report.initialize_structure()
    for file in os.listdir(input_path):
        if os.path.splitext(file)[1] == ".pickle":
            with open(os.path.join(input_path, file),  'rb') as pkl_file:
                test_data = pickle.load(pkl_file)
                ds_name = file.split("_pe_output.pickle")[0]
                report.write_to_sheet(test_data, ds_name)
    report.save_workbook()